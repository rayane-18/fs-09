import openpyxl
import json

# Load the Excel file
excel_file = 'kanji form vocab mine.xlsx'
workbook = openpyxl.load_workbook(excel_file)

# Specify the sheet name
sheet_name = 'Sheet1'  # Change this to the actual sheet name

# Get the worksheet
worksheet = workbook[sheet_name]

# Initialize the quizQuestions dictionary
quiz_questions = {}

# Iterate through rows starting from the second row (assuming the first row contains headers)
for row_num in range(2, worksheet.max_row + 1):
    # Get question and correct answer from columns A and B
    question = worksheet.cell(row=row_num, column=1).value
    correct_answer = worksheet.cell(row=row_num, column=2).value

    # Initialize options list with the correct answer
    options = [correct_answer]

    # Iterate through other columns (C, D, E, ...) to get options
    for col_num in range(3, worksheet.max_column + 1):
        option = worksheet.cell(row=row_num, column=col_num).value
        if option:
            options.append(option)

    # Randomize the order of options
    import random
    random.shuffle(options)

    # Get the category from the last column (assuming it's in column Z)
    category = worksheet.cell(row=row_num, column=26).value or 'Other'

    # Check if the category exists in quizQuestions, otherwise create it
    if category not in quiz_questions:
        quiz_questions[category] = []

    # Add the question to the category
    quiz_questions[category].append({
        'question': question,
        'options': options,
        'correctAnswer': correct_answer,
    })

# Save the quizQuestions dictionary to a JSON file
output_file_path = 'quiz_questions.json'
with open(output_file_path, 'w') as output_file:
    json.dump(quiz_questions, output_file, indent=2)

print(f'Quiz questions generated and saved to {output_file_path}')
